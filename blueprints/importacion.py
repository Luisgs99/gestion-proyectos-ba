import io
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename
import pandas as pd
from flask import Blueprint, render_template, request, redirect, url_for, session, flash, send_file, current_app
from database import query, execute
from helpers.auth import editor_required, login_required

bp = Blueprint('importacion', __name__)


@bp.route('/importar', methods=['GET', 'POST'])
@editor_required
def importar():
    programas = query("SELECT * FROM programas WHERE activo=1 ORDER BY id")
    agentes   = query("SELECT * FROM users WHERE rol='agente' AND activo=1 ORDER BY apellido")

    if request.method == 'POST':
        file       = request.files.get('archivo')
        programa_id= request.form.get('programa_id')
        if not file or not programa_id:
            flash('Seleccioná un archivo y un programa.', 'danger')
            return redirect(url_for('importacion.importar'))

        filename = secure_filename(file.filename)
        filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        try:
            df = pd.read_excel(filepath, engine='openpyxl')
            df.columns = [str(c).strip().lower().replace(' ', '_') for c in df.columns]
            importados, errores, log = 0, 0, []

            hitos = query("SELECT * FROM hitos WHERE programa_id=? ORDER BY orden", (programa_id,))

            for i, row in df.iterrows():
                try:
                    nombre = str(row.get('nombre', row.get('proyecto', f'Proyecto {i+1}'))).strip()
                    if not nombre or nombre == 'nan':
                        continue
                    beneficiario = str(row.get('beneficiario', row.get('institucion', ''))).strip()
                    if beneficiario == 'nan': beneficiario = ''
                    adoptante = str(row.get('adoptante', row.get('empresa', ''))).strip()
                    if adoptante == 'nan': adoptante = ''
                    anr = 0
                    anr_raw = row.get('anr', row.get('monto', row.get('anr_monto', 0)))
                    try:
                        anr = float(str(anr_raw).replace('.', '').replace(',', '.') or 0)
                    except Exception:
                        anr = 0
                    estado_raw = str(row.get('estado', 'activo')).lower().strip()
                    estado_map = {
                        'activo': 'activo', 'finalizado': 'finalizado',
                        'suspendido': 'suspendido',
                        'en evaluación': 'en_evaluacion', 'en_evaluacion': 'en_evaluacion',
                    }
                    estado   = estado_map.get(estado_raw, 'activo')
                    municipio= str(row.get('municipio', '')).strip()
                    if municipio == 'nan': municipio = ''
                    codigo = str(row.get('codigo', row.get('expediente', ''))).strip()
                    if codigo == 'nan': codigo = ''

                    pid = execute("""
                        INSERT INTO proyectos (programa_id, nombre, codigo, beneficiario, adoptante,
                            anr_monto, estado, municipio, agente_id)
                        VALUES (?,?,?,?,?,?,?,?,?)
                    """, (programa_id, nombre, codigo, beneficiario, adoptante,
                          anr, estado, municipio, session['user_id']))

                    for h in hitos:
                        col_hito   = h['nombre'].lower().replace(' ', '_')
                        avance_val = row.get(col_hito, 'pendiente')
                        avance_str = str(avance_val).lower().strip() if avance_val else 'pendiente'
                        estado_hito= 'pendiente'
                        if avance_str in ['completado', 'completo', 'si', 'sí', '1', 'true']:
                            estado_hito = 'completado'
                        elif avance_str in ['en_proceso', 'en proceso', 'proceso']:
                            estado_hito = 'en_proceso'
                        execute("INSERT OR IGNORE INTO avances_hitos (proyecto_id, hito_id, estado, registrado_por) VALUES (?,?,?,?)",
                                (pid, h['id'], estado_hito, session['user_id']))
                    importados += 1
                except Exception as e:
                    errores += 1
                    log.append(f"Fila {i+2}: {str(e)}")

            execute("""INSERT INTO importaciones (programa_id, filename, registros_importados,
                                                  registros_error, log_detalle, importado_por)
                       VALUES (?,?,?,?,?,?)""",
                    (programa_id, filename, importados, errores, '\n'.join(log), session['user_id']))
            flash(f'Importación completada: {importados} proyectos importados, {errores} errores.',
                  'success' if errores == 0 else 'warning')
        except Exception as e:
            flash(f'Error al procesar el archivo: {str(e)}', 'danger')
        return redirect(url_for('importacion.importar'))

    historial = query("""
        SELECT i.*, pr.nombre as programa_nombre, u.nombre as user_nombre, u.apellido as user_apellido
        FROM importaciones i
        LEFT JOIN programas pr ON i.programa_id = pr.id
        LEFT JOIN users u ON i.importado_por = u.id
        ORDER BY i.created_at DESC LIMIT 20
    """)
    return render_template('programs/importar.html', programas=programas, agentes=agentes, historial=historial)


@bp.route('/importar/plantilla/<int:prog_id>')
@login_required
def plantilla(prog_id):
    programa = query("SELECT * FROM programas WHERE id=?", (prog_id,), one=True)
    hitos    = query("SELECT * FROM hitos WHERE programa_id=? ORDER BY orden", (prog_id,))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Plantilla {programa['codigo']}"
    headers  = ['nombre', 'codigo', 'beneficiario', 'adoptante', 'anr_monto',
                 'estado', 'municipio', 'localidad', 'area_tematica']
    headers += [h['nombre'].lower().replace(' ', '_') for h in hitos]
    header_fill = PatternFill(start_color="006AC1", end_color="006AC1", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = max(15, len(h) + 2)
    ws.append(['Ejemplo: Proyecto IA', 'EXP-001', 'UNLP', 'TechCorp SA', 500000, 'activo', 'La Plata', 'La Plata', 'IA/ML'])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name=f"plantilla_{programa['codigo']}.xlsx",
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
